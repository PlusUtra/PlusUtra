import os
import datetime
import time
import pyautogui as py
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from selenium import webdriver
from selenium.webdriver.edge.service import Service
import psutil  # Para manipular processos
import shutil  # Para mover e deletar arquivos

# Função para procurar o arquivo na pasta
def procurar_arquivo(nome_arquivo, caminho_pasta):
    for root, dirs, files in os.walk(caminho_pasta):
        for file in files:
            if file == nome_arquivo:
                return os.path.join(root, file)
    return None

def encontrar_linhas_dentro_do_intervalo(ws):
    hoje = datetime.datetime.now()

    # Subtrair uma hora da hora atual para pegar a hora anterior
    hora_anterior = hoje - datetime.timedelta(hours=1)

    # Definir o intervalo de tempo como sendo a hora anterior
    hora_atual = hora_anterior.hour
    minuto_atual = hora_anterior.minute

    # Definir o início e fim do intervalo (arredondado para a hora anterior)
    data_busca_inicio = hora_anterior.replace(hour=hora_atual, minute=0, second=0, microsecond=0)
    data_busca_fim = data_busca_inicio.replace(hour=hora_atual, minute=59, second=59, microsecond=999999)

    print(f"Intervalo de busca: {data_busca_inicio} - {data_busca_fim}")

    linhas_para_manter = []
    for linha in range(2, ws.max_row + 1):  # Começa em 2 para ignorar cabeçalhos
        valor_c = ws[f'C{linha}'].value  # Coluna C

        if isinstance(valor_c, datetime.datetime):
            if data_busca_inicio <= valor_c <= data_busca_fim:
                linhas_para_manter.append(linha)
        elif isinstance(valor_c, datetime.time):
            valor_c = datetime.datetime.combine(hoje.date(), valor_c)
            if data_busca_inicio <= valor_c <= data_busca_fim:
                linhas_para_manter.append(linha)
        else:
            try:
                valor_c = datetime.datetime.strptime(valor_c, "%d/%m/%Y %H:%M:%S")
                if data_busca_inicio <= valor_c <= data_busca_fim:
                    linhas_para_manter.append(linha)
            except ValueError:
                continue

    return linhas_para_manter

# Função para criar uma nova aba com as linhas dentro do intervalo e excluir a aba original
def transferir_e_excluir_aba(nome_arquivo, caminho_pasta):
    caminho_arquivo = procurar_arquivo(nome_arquivo, caminho_pasta)
    if not caminho_arquivo:
        print(f"Arquivo {nome_arquivo} não encontrado na pasta especificada.")
        return

    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    # Procura as linhas que atendem ao intervalo
    linhas_para_manter = encontrar_linhas_dentro_do_intervalo(ws)

    if linhas_para_manter:
        # Criar uma nova aba para os dados dentro do intervalo
        nova_aba = wb.create_sheet("Form11")

        # Estilo para as colunas B e C (data e hora)
        datetime_style = NamedStyle(name="datetime_style", number_format="MM/DD/YY HH:MM:SS")

        # Copiar cabeçalhos da coluna A até AH
        for col in range(1, 35):  # A até AH (colunas 1 até 35)
            nova_aba.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)

        # Copiar os dados dentro do intervalo para a nova aba
        for i, linha in enumerate(linhas_para_manter, start=2):  # Começa em 2 para preservar cabeçalhos
            for col in range(1, ws.max_column + 1):
                nova_aba.cell(row=i, column=col, value=ws.cell(row=linha, column=col).value)

                # Formatar colunas B e C
                if col == 2 or col == 3:  # Coluna B e C
                    nova_aba.cell(row=i, column=col).style = datetime_style

        # Excluir a aba original
        wb.remove(ws)

        # Garantir que a aba criada seja renomeada para "Form1"
        if "Form11" in wb.sheetnames:
            form11_aba = wb["Form11"]
            form11_aba.title = "Form1"
            print("A aba foi renomeada de 'Form11' para 'Form1'.")

            # Ajustar a largura das colunas de A até AH para 20px
            from openpyxl.utils import get_column_letter
            for col in range(1, 35):  # A até AH (colunas 1 até 35)
                coluna = get_column_letter(col)  # Converte o número da coluna para a letra
                form11_aba.column_dimensions[coluna].width = 20  # Definindo a largura para 20
                print(f"Largura da coluna {coluna} ajustada para 20px.")

        # Salvar o arquivo após as alterações
        wb.save(caminho_arquivo)
        print("Os dados dentro do intervalo foram transferidos para a nova aba 'Form1' e a aba original foi excluída.")
    else:
        print("Não foi encontrado nenhum dado dentro do intervalo.")

        # Se não encontrar dados dentro do intervalo, excluir todas as linhas da planilha
        for linha in range(2, ws.max_row + 1):
            ws.delete_rows(2)  # Excluir a linha 2 em diante

        wb.save(caminho_arquivo)
        print("Todas as linhas foram excluídas da planilha.")


# Função para fechar todos os processos do Excel sem salvar
def fechar_excel_sem_salvar():
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if 'excel' in proc.info['name'].lower():
                proc.terminate()  # Fecha o processo do Excel
                print(f"Processo Excel com PID {proc.info['pid']} encerrado.")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass


# Função para iniciar o navegador Edge com Selenium e realizar os cliques
def iniciar_edge():
    # Caminho onde o EdgeDriver está localizado
    caminho_edge_driver = "C:/Users/oleroyme/OneDrive - DPDHL/Desktop/Projeto OMS/msedgedriver.exe"  # Substitua com o caminho correto
    
    # Inicializar o WebDriver com o novo método utilizando 'Service'
    service = Service(caminho_edge_driver)
    driver = webdriver.Edge(service=service)
    
    # Acessar o link desejado
    driver.get("https://dpdhl-my.sharepoint.com/:x:/r/personal/joao_barrosn_dhl_com/_layouts/15/Doc.aspx?sourcedoc=%7B403D0A57-AC80-4EE2-8940-83CCEAD4E15C%7D&file=Atividades%20Manuais%20LMS%20-%20Leroy%20Merlin.xlsx&action=default&mobileredirect=true")  # Substitua pela URL correta
    print("Microsoft Edge foi iniciado!")

    # Maximizar a janela do navegador
    driver.maximize_window()
    print("Janela maximizada!")

    time.sleep(20)
    py.click(x=60, y=255)
    time.sleep(2)
    py.click(x=332, y=532)
    time.sleep(2)
    py.click(x=552, y=578)
    time.sleep(2)
    
    return driver
    
    # Fechar Excel sem salvar antes de confirmar
    fechar_excel_sem_salvar()

# Função para mover o arquivo para a pasta de destino
def mover_arquivo(nome_arquivo, origem, destino):
    caminho_arquivo_origem = os.path.join(origem, nome_arquivo)
    caminho_arquivo_destino = os.path.join(destino, nome_arquivo)

    if os.path.exists(caminho_arquivo_destino):
        print(f"Arquivo com o nome '{nome_arquivo}' já existe na pasta de destino. Deletando o arquivo antigo...")
        os.remove(caminho_arquivo_destino)  # Deletar o arquivo existente

    # Mover o arquivo para a pasta de destino
    shutil.move(caminho_arquivo_origem, caminho_arquivo_destino)
    print(f"Arquivo movido para: {caminho_arquivo_destino}")

# Função principal para garantir que o processo ocorra na sequência desejada
# Função principal de execução da automação
def executar_automacao():
    # Executar processo do Selenium
    driver = iniciar_edge()
    time.sleep(10)  # Aguarda um pouco para a página carregar
    driver.quit()  # Fecha o navegador após o processo do Selenium

    # Agora, procura o arquivo na pasta de Downloads
    caminho_pasta = "C:/Users/oleroyme/Downloads"
    nome_arquivo = "Atividades Manuais LMS - Leroy Merlin.xlsx"  # Nome fixo do arquivo Excel

    caminho_arquivo = procurar_arquivo(nome_arquivo, caminho_pasta)

    if caminho_arquivo:
        print(f"Arquivo encontrado: {caminho_arquivo}")
        
        # Renomear o arquivo antes de movê-lo
        novo_nome_arquivo = "Atividades Manuais LMS.xlsx"  # Nome desejado para o arquivo
        novo_caminho_arquivo = os.path.join(caminho_pasta, novo_nome_arquivo)

        # Renomeando o arquivo
        os.rename(caminho_arquivo, novo_caminho_arquivo)
        print(f"Arquivo renomeado para: {novo_nome_arquivo}")

        # Aplica a lógica de corte de horário no arquivo (se necessário)
        transferir_e_excluir_aba(novo_nome_arquivo, caminho_pasta)
        
        # Mover o arquivo renomeado para a pasta de destino
        mover_arquivo(novo_nome_arquivo, caminho_pasta, "C:/Users/oleroyme/OneDrive - DPDHL/Desktop/Projeto OMS")
    else:
        print(f"O arquivo '{nome_arquivo}' não foi encontrado na pasta de Downloads.")

# Chamando a função principal para executar a automação
executar_automacao()

#-----------------AUTOMAÇÃO VBA-------------------------------#

import time
import psutil
import win32com.client  # Importação da biblioteca para interagir com o Excel

def focar_janela_excel(caminho_arquivo):
    """
    Função para abrir o Excel, maximizar a janela e focar nela.
    
    :param caminho_arquivo: Caminho completo para o arquivo Excel (.xlsm)
    """
    # Iniciar o Excel com a interface gráfica visível
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True  # Isso faz o Excel aparecer na tela

    # Abrir o arquivo do Excel
    workbook = excel.Workbooks.Open(caminho_arquivo)

    # Maximizar a janela do Excel
    excel.WindowState = 2  # 2 corresponde ao estado maximizado

    # Focar na janela com o nome específico do arquivo
    for window in excel.Windows:
        if window.Caption == "Atualizar LMS.xlsm":  # Verifica se o título da janela corresponde ao nome do arquivo
            window.WindowState = 2  # Maximiza a janela
            window.Activate()  # Ativa a janela para colocar o foco nela
            time.sleep(1)
            py.click(x=681, y=666)

    print(f"Excel focado e janela maximizada com sucesso para o arquivo '{caminho_arquivo}'.")

def fechar_excel_aberto():
    """
    Função para fechar todas as instâncias abertas do Excel, incluindo aquelas fora do controle do script.
    """
    # Fechar todas as instâncias de Excel em execução
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # Verifica se o processo é do Excel
            if 'excel' in proc.info['name'].lower():
                proc.terminate()  # Encerra o processo do Excel
                print(f"Processo do Excel com PID {proc.info['pid']} encerrado.")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

    print("Todos os processos do Excel foram fechados.")

# Caminho do arquivo
caminho_arquivo = r"C:\Users\oleroyme\OneDrive - DPDHL\Desktop\Projeto OMS\Atualizar LMS.xlsm"

# Abrir o Excel, maximizar e focar na janela
focar_janela_excel(caminho_arquivo)

# Aguardar 45 segundos
time.sleep(45)

# Fechar todos os arquivos e processos abertos no Excel após 45 segundos
fechar_excel_aberto()
    
#----------- Corte Base SAP ---------- #

import os
import shutil

def mover_arquivo_para_oms(diretorio_origem, diretorio_destino):
    try:
        # Lista os arquivos na pasta de origem (Downloads)
        arquivos = os.listdir(diretorio_origem)
        
        # Filtra arquivos com extensão '.xlsx' (ou outra que preferir)
        arquivos_xlsx = [arquivo for arquivo in arquivos if arquivo.endswith('.xlsx')]

        if not arquivos_xlsx:
            print("Nenhum arquivo .xlsx encontrado na pasta Downloads.")
            return

        # Pega o arquivo mais recente da lista
        arquivo_mais_recente = max(arquivos_xlsx, key=lambda f: os.path.getmtime(os.path.join(diretorio_origem, f)))
        
        # Caminhos completos dos arquivos
        caminho_origem_arquivo = os.path.join(diretorio_origem, arquivo_mais_recente)
        caminho_destino_arquivo = os.path.join(diretorio_destino, arquivo_mais_recente)

        # Verifica se o arquivo já existe no diretório de destino
        if os.path.exists(caminho_destino_arquivo):
            os.remove(caminho_destino_arquivo)  # Remove o arquivo existente no destino
            print(f"O arquivo existente foi removido de {diretorio_destino}.")
        
        # Move o arquivo para o diretório de destino
        shutil.move(caminho_origem_arquivo, caminho_destino_arquivo)
        print(f"O arquivo '{arquivo_mais_recente}' foi movido para {diretorio_destino}.")
    
    except Exception as e:
        print(f"Erro ao mover o arquivo: {e}")

# Testando a função
diretorio_origem = r"C:\Users\oleroyme\Downloads"
diretorio_destino = r"C:\Users\oleroyme\OneDrive - DPDHL\Desktop\Projeto OMS"

mover_arquivo_para_oms(diretorio_origem, diretorio_destino)


#------------- Sharpoint -----------#

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui as py

def iniciar_sharepoint():
    """
    Função para iniciar o navegador Microsoft Edge e interagir com o SharePoint.
    """
    # Caminho onde o EdgeDriver está localizado
    caminho_edge_driver = "C:/Users/oleroyme/OneDrive - DPDHL/Desktop/Projeto OMS/msedgedriver.exe"  # Substitua com o caminho correto

    # Inicializar o WebDriver com o novo método utilizando 'Service'
    service = Service(caminho_edge_driver)
    driver = webdriver.Edge(service=service)

    # Acessar o link desejado
    url_sharepoint = "https://dpdhl.sharepoint.com/teams/DataModeler/Shared%20Documents/Forms/AllItems.aspx?csf=1&web=1&e=wbM0IF&ovuser=cd99fef8%2D1cd3%2D4a2a%2D9bdf%2D15531181d65e%2Cjonatan%2Emilfont%40dhl%2Ecom&CID=3cac72a1%2D0011%2D8000%2Df640%2Dd7030655b014&cidOR=SPO&FolderCTID=0x01200025AEFFD1A05D694EB7B6BAB9629F094A&id=%2Fteams%2FDataModeler%2FShared%20Documents%2FLeroy%20Merlin%20%2D%20Cajamar&OR=Teams%2DHL&CT=1737026219348&clickparams=eyJBcHBOYW1lIjoiVGVhbXMtRGVza3RvcCIsIkFwcFZlcnNpb24iOiI0OS8yNDEyMDEwMDIxNyIsIkhhc0ZlZGVyYXRlZFVzZXIiOmZhbHNlfQ%3D%3D"  # Substitua pela URL correta
    driver.get(url_sharepoint)
    print("Microsoft Edge foi iniciado!")

    # Maximizar a janela do navegador
    driver.maximize_window()
    print("Janela maximizada!")

    try:
        # ---------- Primeira interação (Clique no botão dentro do div) ----------
        primeiro_elemento = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='appRoot']/div[1]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div/div/div/div/div/div[1]/div[2]/button/span"))
        )
        primeiro_elemento.click()
        print("Primeiro clique realizado com sucesso!")

        # ---------- Segunda interação (Clique no botão 'Files' dentro do Upload) ----------
        segundo_elemento = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//span[text()='Files']"))
        )
        segundo_elemento.click()
        print("Segundo clique realizado com sucesso!")

        # ---------- Selecionando arquivos ----------
        time.sleep(5)

        # Clique para abrir o endereço do arquivo
        py.click(x=508, y=83)  # Substitua com as coordenadas do primeiro item
        time.sleep(2)
        py.write('C:/Users/oleroyme/OneDrive - DPDHL/Desktop/Projeto OMS')
        py.press('Enter')
        time.sleep(3)

        # Clique no primeiro arquivo
        py.click(x=440, y=385)  # Substitua com as coordenadas do segundo item
        print("Clique realizado no primeiro arquivo!")

        # Pressionar a tecla Ctrl para selecionar múltiplos arquivos
        py.keyDown('ctrl')  # Segura a tecla Ctrl
        print("Ctrl pressionado!")
        time.sleep(2)

        # Clique no segundo arquivo com Ctrl pressionado
        py.click(x=439, y=474)  # Substitua com as coordenadas do segundo item
        print("Clique realizado no segundo arquivo com Ctrl pressionado!")

        # Liberar a tecla Ctrl
        py.keyUp('ctrl')  # Libera a tecla Ctrl
        print("Ctrl liberado!")

        time.sleep(3)
        py.press('Enter')
        time.sleep(10)

        # ---------- Clique no botão 'Replace all' ----------
        replace_all_elemento = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//span[text()='Replace all']"))
        )
        replace_all_elemento.click()
        print("Clique no 'Replace all' realizado com sucesso!")

    except Exception as e:
        print(f"Erro ao tentar clicar nos elementos: {e}")

    # Aguardar mais tempo, se necessário
    time.sleep(15)
    
    return driver

# Exemplo de uso
if __name__ == "__main__":
    driver = iniciar_sharepoint()

    # Adicione código para interagir com o SharePoint ou aguardar o carregamento da página, se necessário.
    # O código de fechamento do driver foi removido conforme solicitado.

py.click(x=781, y=642)

#-----------------AUTOMAÇÃO PBI-------------------------------#

import os
import subprocess

def abrir_arquivo_especifico(caminho_arquivo):
    """
    Função para verificar se o arquivo existe e abrir o arquivo especificado.
    """
    if os.path.exists(caminho_arquivo):
        print(f"Arquivo encontrado: {caminho_arquivo}")
        try:
            # Tenta abrir o arquivo com o programa padrão associado
            subprocess.Popen([caminho_arquivo], shell=True)
            print("Arquivo aberto com sucesso!")
        except Exception as e:
            print(f"Erro ao tentar abrir o arquivo: {e}")
    else:
        print(f"O arquivo '{caminho_arquivo}' não foi encontrado no caminho especificado.")

# Caminho completo para o arquivo .pbix
caminho_arquivo = r"C:\Users\oleroyme\OneDrive - DPDHL\Desktop\Data Modeler - Leroy Merlin Cajamar 22.pbix"

# Chamar a função para abrir o arquivo
abrir_arquivo_especifico(caminho_arquivo)


time.sleep(63) #Abrindo PBI

py.click(x=1156, y=138) #Atualizando PBI
time.sleep(65) #Aguardando Atualizar
py.click(x=1239, y=732) #Fechando diálogo de atualização
time.sleep(5) #Aguardando 5 Segundos
py.click(x=1254, y=574)# clique na tabela
time.sleep(1) # Aguarda 1 segundo
py.click(x=1272, y=487) #Clicando nos 3 pontinhos
time.sleep(2) #Aguardando 2 Segundos
py.click(x=1340, y=513) #Clicando em exportar
time.sleep(10) #Aguardando 10 Segundos
py.click(x=492, y=80) #Clicando no endereço de salvamento
time.sleep(1) #Aguardando 1 Segundo
py.write('C:/Users/oleroyme/OneDrive - DPDHL/Desktop/Projeto OMS') # Escrevendo o caminho para salvar
time.sleep(2) # Aguardando 1 Segundo
py.press('Enter')
time.sleep(4)
py.click(x=528, y=492) #Clicando no nome do arquivo
time.sleep(3) # Aguarda 1 segundo
py.write('Mload') # Renomeando arquivo
time.sleep(3)
py.press('Enter')   
time.sleep(2)
py.press('tab')
time.sleep(1)
py.press('Enter')

# ----------------DHL Link Automação------------#


import time
import datetime
import pyautogui as py
from selenium import webdriver
from selenium.webdriver.edge.service import Service

def iniciar_dhllink():
    """
    Função para iniciar o Microsoft Edge, acessar o link desejado no DHL e realizar uma sequência de interações.
    """
    try:
        # Caminho onde o EdgeDriver está localizado
        caminho_edge_driver = "C:/Users/oleroyme/OneDrive - DPDHL/Desktop/Projeto OMS/msedgedriver.exe"  # Substitua com o caminho correto
        
        # Inicializar o WebDriver com o novo método utilizando 'Service'
        service = Service(caminho_edge_driver)
        driver = webdriver.Edge(service=service)
        
        # Acessar o link desejado
        link_dhl = "https://link-cc.dhl.com/"  # Substitua pela URL correta
        driver.get(link_dhl)
        print("Microsoft Edge foi iniciado!")

        # Maximizar a janela do navegador
        driver.maximize_window()
        print("Janela maximizada!")
        
        # Esperar 30 segundos para garantir que a página carregue completamente
        time.sleep(30)
        
        # Interações com a interface gráfica
        try:
            py.click(x=488, y=198)  # Primeiro clique
            time.sleep(5)
            
            py.click(x=805, y=686)  # Segundo clique
            time.sleep(5)
            
            py.click(x=749, y=73)  # Terceiro clique
            py.write('Mload')  # Escrevendo 'Mload'
            time.sleep(2)
            
            py.click(x=681, y=192)  # Quarto clique
            time.sleep(4)
            
            py.press('Enter')
            
            time.sleep(15)
            
            print("Automação concluída com sucesso!")
        except Exception as e:
            print(f"Erro ao realizar as interações com a interface gráfica: {e}")
        
        return driver  # Retorna o driver após as interações
    except Exception as e:
        print(f"Erro ao iniciar o Microsoft Edge: {e}")
        return None

# Função para esperar até a próxima hora cheia
def esperar_nova_hora():
    """
    Espera até a próxima hora cheia para executar a próxima tarefa.
    """
    while True:
        now = datetime.datetime.now()
        # Calcula quantos segundos faltam até a próxima hora cheia
        seconds_until_next_hour = 3600 - (now.minute * 60 + now.second)
        
        print(f"Aguardando {seconds_until_next_hour} segundos até a próxima hora.")
        time.sleep(seconds_until_next_hour)

if __name__ == "__main__":
    # Iniciar o processo no Microsoft Edge
    driver = iniciar_dhllink()

    if driver:
        # Finalizar o driver após concluir as interações
        driver.quit()
        print("Driver do Microsoft Edge fechado com sucesso.")
    else:
        print("Erro ao iniciar o Microsoft Edge. Não foi possível continuar o processo.")

    # Esperar até a próxima hora cheia, caso necessário
    esperar_nova_hora()

import psutil

def fechar_navegadores():
    """
    Função para fechar todos os navegadores que estão abertos (Chrome, Edge, Firefox).
    """
    # Lista de nomes de processos dos navegadores
    navegadores = ["chrome", "edge", "firefox"]

    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if any(navegador in proc.info['name'].lower() for navegador in navegadores):
                proc.terminate()  # Tenta encerrar o processo
                print(f"Processo {proc.info['name']} com PID {proc.info['pid']} encerrado.")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

# Exemplo de uso no final do seu código
if __name__ == "__main__":
    fechar_navegadores()
